from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file, jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_socketio import SocketIO, emit
from flask_mail import Mail, Message
from datetime import datetime, timedelta
from functools import wraps
from werkzeug.utils import secure_filename
import random
import string
import os
import qrcode
from io import BytesIO
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import distinct




app = Flask(__name__)
app.config['SECRET_KEY'] = 'aai-ims-secret-key-2026'
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:///assets.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# ===================== MAIL CONFIG =====================
# Replace with your Gmail credentials
# For Gmail: enable 2FA → go to myaccount.google.com/apppasswords → generate app password
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'your-atharvg0705@gmail.com'       # <-- CHANGE THIS
app.config['MAIL_PASSWORD'] = 'evel ufdx datu fuep'     # <-- CHANGE THIS (app password, not account password)
app.config['MAIL_DEFAULT_SENDER'] = 'AAI IMS <your-atharvg0705@gmail.com>'
app.config['MAIL_SUPPRESS_SEND'] = False  # Set True during dev/testing to skip actual send

# File upload configuration
UPLOAD_FOLDER = 'static/uploads/documents'
ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg'}
MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_FILE_SIZE

# Create upload folder if it doesn't exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


db = SQLAlchemy(app)
socketio = SocketIO(app, cors_allowed_origins="*")
mail = Mail(app)



# ===================== MODELS =====================



class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password = db.Column(db.String(120), nullable=False)
    role = db.Column(db.String(20), nullable=False)
    employee_id = db.Column(db.Integer, db.ForeignKey('employee.id'), nullable=True)



class Employee(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.String(50), unique=True, nullable=False)
    name = db.Column(db.String(120), nullable=False)
    email = db.Column(db.String(120))
    mobile = db.Column(db.String(20))
    department = db.Column(db.String(80), nullable=False)
    designation = db.Column(db.String(80))
    is_available = db.Column(db.Boolean, default=True)
    current_workload = db.Column(db.Integer, default=0)
    max_tickets = db.Column(db.Integer, default=5)
    employee_type = db.Column(db.String(20), default='staff') 


class Asset(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    asset_name = db.Column(db.String(120), nullable=False)
    serial_number = db.Column(db.String(120), nullable=False, unique=True)
    asset_tag = db.Column(db.String(120))
    category = db.Column(db.String(80), nullable=False)
    location = db.Column(db.String(120), nullable=False)
    site = db.Column(db.String(120))
    status = db.Column(db.String(50), nullable=False, default='In Store')
    barcode = db.Column(db.String(120))
    service_tag = db.Column(db.String(120))
    model = db.Column(db.String(120))
    manufacturer = db.Column(db.String(120))
    business_impact = db.Column(db.String(50))
    department = db.Column(db.String(100))
    os = db.Column(db.String(100))
    service_pack = db.Column(db.String(50))
    ram_gb = db.Column(db.String(50))
    virtual_memory_gb = db.Column(db.String(50))
    processor_info = db.Column(db.String(200))
    processor_manufacturer = db.Column(db.String(100))
    cpu_speed_ghz = db.Column(db.String(50))
    number_of_cores = db.Column(db.Integer)
    hdd_model = db.Column(db.String(100))
    hdd_serial = db.Column(db.String(100))
    hdd_capacity_gb = db.Column(db.String(50))
    ip_address = db.Column(db.String(50))
    mac_address = db.Column(db.String(50))
    nic = db.Column(db.String(100))
    default_gateway = db.Column(db.String(50))
    network = db.Column(db.String(100))
    subnet_mask = db.Column(db.String(50))
    dhcp_enabled = db.Column(db.Boolean, default=False)
    dhcp_server = db.Column(db.String(50))
    vendor_name = db.Column(db.String(120))
    purchase_cost = db.Column(db.Float, default=0.0)
    acquisition_date = db.Column(db.Date)
    expiry_date = db.Column(db.Date)
    warranty_expiry = db.Column(db.String(20))
    monitor = db.Column(db.String(200))
    comments = db.Column(db.Text)
    assigned_employee_id = db.Column(db.Integer, db.ForeignKey('employee.id'), nullable=True)
    assigned_to = db.Column(db.String(120))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    assignments = db.relationship('AssetAssignment', backref='asset', lazy=True)
    reports = db.relationship('AssetReport', backref='asset', lazy=True)



class AssetAssignment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    asset_id = db.Column(db.Integer, db.ForeignKey('asset.id'), nullable=False)
    employee_id = db.Column(db.Integer, db.ForeignKey('employee.id'), nullable=False)
    assigned_date = db.Column(db.DateTime, default=datetime.utcnow)
    returned_date = db.Column(db.DateTime, nullable=True)
    notes = db.Column(db.Text, default='Issued to employee')



class AssetReport(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    asset_id = db.Column(db.Integer, db.ForeignKey('asset.id'), nullable=False)
    employee_id = db.Column(db.Integer, db.ForeignKey('employee.id'), nullable=False)
    report_type = db.Column(db.String(50))
    message = db.Column(db.Text, nullable=False)
    status = db.Column(db.String(20), default='pending')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    reporter = db.relationship('Employee', backref='reports')



class Ticket(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    ticket_number = db.Column(db.String(20), unique=True, nullable=False)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text, nullable=False)
    priority = db.Column(db.String(20), default='Medium')
    category = db.Column(db.String(50))
    asset_id = db.Column(db.Integer, db.ForeignKey('asset.id'))
    raised_by_employee_id = db.Column(db.Integer, db.ForeignKey('employee.id'), nullable=False)
    assigned_to_employee_id = db.Column(db.Integer, db.ForeignKey('employee.id'))
    status = db.Column(db.String(30), default='Open')
    raised_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    assigned_at = db.Column(db.DateTime)
    resolved_at = db.Column(db.DateTime)
    closed_at = db.Column(db.DateTime)
    reminder_sent = db.Column(db.Boolean, default=False)
    reminder_sent_at = db.Column(db.DateTime)
    resolution_notes = db.Column(db.Text)
    raised_by = db.relationship('Employee', foreign_keys=[raised_by_employee_id], backref='tickets_raised')
    assigned_to = db.relationship('Employee', foreign_keys=[assigned_to_employee_id], backref='tickets_assigned')
    asset = db.relationship('Asset', backref='tickets')

    def is_overdue(self):
        if self.status in ['Resolved', 'Closed']:
            return False
        time_elapsed = datetime.utcnow() - self.raised_at
        return time_elapsed > timedelta(hours=2)

    def time_since_raised(self):
        delta = datetime.utcnow() - self.raised_at
        hours = delta.total_seconds() / 3600
        if hours < 1:
            return f"{int(delta.total_seconds() / 60)} minutes ago"
        elif hours < 24:
            return f"{int(hours)} hours ago"
        else:
            return f"{int(hours / 24)} days ago"



class AssetDocument(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    asset_id = db.Column(db.Integer, db.ForeignKey('asset.id'), nullable=False)
    document_name = db.Column(db.String(200), nullable=False)
    document_type = db.Column(db.String(100))
    file_path = db.Column(db.String(300), nullable=False)
    remarks = db.Column(db.String(500))
    uploaded_by = db.Column(db.Integer, db.ForeignKey('employee.id'))
    uploaded_at = db.Column(db.DateTime, default=datetime.utcnow)
    asset = db.relationship('Asset', backref='documents')
    uploader = db.relationship('Employee', backref='uploaded_documents')



class AuditLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    action = db.Column(db.String(100), nullable=False)
    entity_type = db.Column(db.String(50), nullable=False)
    entity_id = db.Column(db.Integer)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow)
    details = db.Column(db.Text)
    user = db.relationship('User', backref='audit_logs')



# ===================== NEW: NOTIFICATION MODEL =====================

class Notification(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    employee_id = db.Column(db.Integer, db.ForeignKey('employee.id'), nullable=False)
    message = db.Column(db.String(300), nullable=False)
    link = db.Column(db.String(100))
    is_read = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    employee = db.relationship('Employee', backref='notifications')



# ===================== AUTH DECORATORS =====================



def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please login first.', 'danger')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function



def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if session.get('role') != 'admin':
            flash('Admin access required.', 'danger')
            return redirect(url_for('employee_dashboard'))
        return f(*args, **kwargs)
    return decorated_function



# ===================== HELPERS =====================



def log_action(action, entity_type, entity_id=None, details=None):
    try:
        user_id = session.get('user_id')
    except RuntimeError:
        user_id = None
    log = AuditLog(
        user_id=user_id,
        action=action,
        entity_type=entity_type,
        entity_id=entity_id,
        details=details
    )
    db.session.add(log)



def generate_asset_qr(asset):
    qr_dir = os.path.join('static', 'qr_codes')
    os.makedirs(qr_dir, exist_ok=True)
    data = f"AAI-IMS-ASSET:{asset.id}:{asset.asset_name}:{asset.serial_number}"
    img = qrcode.make(data)
    filename = f"asset_{asset.id}.png"
    filepath = os.path.join(qr_dir, filename)
    img.save(filepath)
    return f"qr_codes/{filename}"



def compute_stats():
    total = Asset.query.count()
    in_store = Asset.query.filter_by(status='In Store').count()
    deployed = Asset.query.filter_by(status='Deployed').count()
    maintenance = Asset.query.filter_by(status='In Maintenance').count()
    in_store_val = db.session.query(db.func.sum(Asset.purchase_cost))        .filter(Asset.status == 'In Store').scalar() or 0
    deployed_val = db.session.query(db.func.sum(Asset.purchase_cost))        .filter(Asset.status == 'Deployed').scalar() or 0
    return {
        "total_assets": total,
        "in_store": in_store,
        "deployed": deployed,
        "in_maintenance": maintenance,
        "in_store_value": in_store_val,
        "deployed_value": deployed_val
    }



def get_employee_by_name(name):
    if not name:
        return None
    return Employee.query.filter(Employee.name == name).first()



def generate_ticket_number():
    year = datetime.now().year
    random_num = ''.join(random.choices(string.digits, k=6))
    return f"TKT-{year}-{random_num}"



# ===================== NEW: EMAIL HELPER =====================

def send_email(to, subject, html_body):
    """Send an email. Silently fails if mail not configured."""
    if not to:
        return
    try:
        msg = Message(subject=subject, recipients=[to])
        msg.html = html_body
        msg.body = html_body.replace('<br>', '\n').replace('<b>', '').replace('</b>', '')
        mail.send(msg)
        print(f"✅ Email sent to {to}")
    except Exception as e:
        print(f"⚠️ Email not sent (check MAIL config): {e}")



# ===================== NEW: NOTIFICATION HELPER =====================

# NEW - FIXED
def create_notification(employee_id, message, link=None):
    notif = Notification(
        employee_id=employee_id,
        message=message,
        link=link
    )
    db.session.add(notif)
    socketio.emit('new_notification', {
        'employee_id': employee_id,
        'message': message,
        'link': link
    })   # <-- removed broadcast=True



# ===================== CONTEXT PROCESSOR (bell icon count) =====================

@app.context_processor
def inject_notification_count():
    """Inject unread notification count into ALL templates automatically."""
    emp_id = session.get('employee_id')
    unread_count = 0
    if emp_id:
        unread_count = Notification.query.filter_by(
            employee_id=emp_id, is_read=False
        ).count()
    return dict(unread_notification_count=unread_count)



# ===================== AUTH ROUTES =====================

@app.route('/api/notifications')
@login_required
def api_notifications():
    emp_id = session.get('employee_id')
    notifs = Notification.query.filter_by(employee_id=emp_id)\
             .order_by(Notification.created_at.desc()).limit(20).all()
    return jsonify([{
        'id':         n.id,
        'message':    n.message,
        'link':       n.link or '',
        'is_read':    n.is_read,
        'created_at': n.created_at.strftime('%d %b, %I:%M %p')
    } for n in notifs])

@app.route('/notifications/mark-all-read', methods=['POST'])
@login_required
def mark_all_notifications_read():
    emp_id = session.get('employee_id')
    Notification.query.filter_by(employee_id=emp_id, is_read=False)\
                .update({'is_read': True})
    db.session.commit()
    return jsonify(success=True)


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username, password=password).first()
        if user:
            session['user_id'] = user.id
            session['username'] = user.username
            session['role'] = user.role
            session['employee_id'] = user.employee_id
            flash(f'Welcome {username}!', 'success')
            if user.role == 'admin':
                return redirect(url_for('index'))
            else:
                return redirect(url_for('employee_dashboard'))
        else:
            flash('Invalid credentials!', 'danger')
    return render_template('login.html')



@app.route('/logout')
def logout():
    session.clear()
    flash('Logged out successfully.', 'success')
    return redirect(url_for('login'))



# ===================== ADMIN ROUTES =====================

@app.route('/reset-employee-password/<int:emp_id>', methods=['POST'])
@login_required
@admin_required
def reset_employee_password(emp_id):
    user = User.query.filter_by(employee_id=emp_id).first()
    if not user:
        flash('No login account found for this employee.', 'danger')
        return redirect(url_for('list_employees'))
    new_password = request.form.get('new_password', '').strip()
    if not new_password or len(new_password) < 4:
        flash('Password must be at least 4 characters.', 'danger')
        return redirect(url_for('list_employees'))
    user.password = new_password
    db.session.commit()
    flash(f'Password updated for {user.username}!', 'success')
    return redirect(url_for('list_employees'))


@app.route('/debug-assets')
@login_required
@admin_required
def debug_assets():
    results = db.session.query(
        Asset,
        Employee.name.label('employee_name')
    ).outerjoin(Employee, Asset.assigned_employee_id == Employee.id).all()
    debug_info = []
    for asset, employee_name in results:
        debug_info.append({
            'asset_id': asset.id,
            'asset_name': asset.asset_name,
            'assigned_employee_id': asset.assigned_employee_id,
            'assigned_to': asset.assigned_to,
            'employee_name': employee_name
        })
    return jsonify(debug_info)


@app.route('/')
@login_required
@admin_required
def index():
    stats = compute_stats()
    results = db.session.query(
        Asset,
        Employee.name.label('employee_name')
    ).outerjoin(Employee, Asset.assigned_employee_id == Employee.id)     .order_by(Asset.created_at.desc())     .limit(10)     .all()
    recent_assets = []
    for asset, employee_name in results:
        asset.employee_name = employee_name or 'Unassigned'
        recent_assets.append(asset)
    return render_template('index.html', stats=stats, recent_assets=recent_assets)


@app.route('/assets')
@login_required
@admin_required
def list_assets():
    search = request.args.get('search', '').strip()
    category = request.args.get('category', '')
    status = request.args.get('status', '')
    query = db.session.query(
        Asset,
        Employee.name.label('employee_name')
    ).outerjoin(Employee, Asset.assigned_employee_id == Employee.id)
    if search:
        like = f"%{search}%"
        query = query.filter(
            db.or_(Asset.asset_name.ilike(like),
                   Asset.serial_number.ilike(like))
        )
    if category:
        query = query.filter(Asset.category == category)
    if status:
        query = query.filter(Asset.status == status)
    results = query.order_by(Asset.created_at.desc()).all()
    assets = []
    for asset, employee_name in results:
        asset.employee_name = employee_name or 'Unassigned'
        assets.append(asset)
    return render_template(
        'assets.html',
        assets=assets,
        search_query=search,
        category=category,
        status=status
    )



@app.route('/export-assets')
@login_required
@admin_required
def export_assets():
    assets = Asset.query.order_by(Asset.created_at.desc()).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assets"
    columns = [
        ("Asset Name", "asset_name"),
        ("Serial Number", "serial_number"),
        ("Category", "category"),
        ("Location", "location"),
        ("Status", "status"),
        ("Assigned To", "assigned_to"),
        ("Purchase Cost", "purchase_cost"),
        ("Acquisition Date", "acquisition_date"),
        ("Expiry Date", "expiry_date"),
        ("Vendor", "vendor_name"),
    ]
    for col_num, (header, _) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
    for row_num, asset in enumerate(assets, start=2):
        for col_num, (_, attr) in enumerate(columns, start=1):
            value = getattr(asset, attr, None)
            if isinstance(value, datetime):
                value = value.strftime("%Y-%m-%d")
            ws.cell(row=row_num, column=col_num, value=value)
    for col_num in range(1, len(columns) + 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 20
    export_dir = os.path.join("static", "exports")
    os.makedirs(export_dir, exist_ok=True)
    filepath = os.path.join(export_dir, f"assets_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    wb.save(filepath)
    return send_file(filepath, as_attachment=True)



@app.route('/add-asset', methods=['GET', 'POST'])
@login_required
@admin_required
def add_asset():
    employees = Employee.query.order_by(Employee.name).all()
    if request.method == 'POST':
        try:
            assigned_name = request.form.get('assigned_to') or None
            emp = get_employee_by_name(assigned_name)
            assigned_id = emp.id if emp else None
            acq_date = None
            if request.form.get('acquisition_date'):
                try:
                    acq_date = datetime.strptime(request.form.get('acquisition_date'), '%Y-%m-%d').date()
                except:
                    pass
            exp_date = None
            if request.form.get('expiry_date'):
                try:
                    exp_date = datetime.strptime(request.form.get('expiry_date'), '%Y-%m-%d').date()
                except:
                    pass
            asset = Asset(
                asset_name=request.form['asset_name'],
                serial_number=request.form['serial_number'],
                asset_tag=request.form.get('asset_tag') or None,
                category=request.form['category'],
                location=request.form['location'],
                site=request.form.get('site') or None,
                status=request.form['status'],
                barcode=request.form.get('barcode') or None,
                service_tag=request.form.get('service_tag') or None,
                model=request.form.get('model') or None,
                manufacturer=request.form.get('manufacturer') or None,
                business_impact=request.form.get('business_impact') or None,
                department=request.form.get('department') or None,
                os=request.form.get('os') or None,
                service_pack=request.form.get('service_pack') or None,
                ram_gb=request.form.get('ram_gb') or None,
                virtual_memory_gb=request.form.get('virtual_memory_gb') or None,
                processor_info=request.form.get('processor_info') or None,
                processor_manufacturer=request.form.get('processor_manufacturer') or None,
                cpu_speed_ghz=request.form.get('cpu_speed_ghz') or None,
                number_of_cores=int(request.form.get('number_of_cores')) if request.form.get('number_of_cores') else None,
                hdd_model=request.form.get('hdd_model') or None,
                hdd_serial=request.form.get('hdd_serial') or None,
                hdd_capacity_gb=request.form.get('hdd_capacity_gb') or None,
                ip_address=request.form.get('ip_address') or None,
                mac_address=request.form.get('mac_address') or None,
                nic=request.form.get('nic') or None,
                default_gateway=request.form.get('default_gateway') or None,
                network=request.form.get('network') or None,
                subnet_mask=request.form.get('subnet_mask') or None,
                dhcp_enabled=bool(int(request.form.get('dhcp_enabled', 0))),
                dhcp_server=request.form.get('dhcp_server') or None,
                vendor_name=request.form.get('vendor_name') or None,
                purchase_cost=float(request.form.get('purchase_cost') or 0),
                acquisition_date=acq_date,
                expiry_date=exp_date,
                warranty_expiry=request.form.get('warranty_expiry') or None,
                monitor=request.form.get('monitor') or None,
                comments=request.form.get('comments') or None,
                assigned_employee_id=assigned_id,
                assigned_to=assigned_name
            )
            db.session.add(asset)
            db.session.commit()
            if assigned_id:
                assignment = AssetAssignment(
                    asset_id=asset.id,
                    employee_id=assigned_id,
                    notes=f"Asset assigned to {assigned_name}"
                )
                db.session.add(assignment)
                db.session.commit()
                # NEW: Notify employee about asset assignment
                create_notification(
                    employee_id=assigned_id,
                    message=f"Asset '{asset.asset_name}' ({asset.serial_number}) has been assigned to you.",
                    link=f"/employee-dashboard"
                )
                db.session.commit()
            if 'document' in request.files:
                file = request.files['document']
                if file and file.filename != '' and allowed_file(file.filename):
                    filename = secure_filename(file.filename)
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    filename = f"{asset.id}_{timestamp}_{filename}"
                    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                    file.save(filepath)
                    document = AssetDocument(
                        asset_id=asset.id,
                        document_name=request.form.get('document_name', 'Asset Document'),
                        document_type=request.form.get('document_type', 'Other'),
                        file_path=f'uploads/documents/{filename}',
                        remarks=request.form.get('document_remarks'),
                        uploaded_by=session.get('employee_id')
                    )
                    db.session.add(document)
                    db.session.commit()
                    flash('Asset and document added successfully!', 'success')
                else:
                    if file and file.filename != '':
                        flash('Asset added! (Document skipped - invalid file type)', 'warning')
                    else:
                        flash('Asset added successfully!', 'success')
            else:
                flash('Asset added successfully!', 'success')
            generate_asset_qr(asset)
            log_action(action="CREATE", entity_type="Asset", entity_id=asset.id,
                       details=f"Asset created: {asset.asset_name} ({asset.serial_number})")
            db.session.commit()
            return redirect('/assets')
        except Exception as e:
            db.session.rollback()
            flash(f'Error adding asset: {str(e)}', 'danger')
    return render_template('add_asset.html', employees=employees)



@app.route('/edit-asset/<int:asset_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_asset(asset_id):
    asset = Asset.query.get_or_404(asset_id)
    employees = Employee.query.order_by(Employee.name).all()
    if request.method == 'POST':
        try:
            assigned_name = request.form.get('assigned_to') or None
            emp = get_employee_by_name(assigned_name)
            assigned_id = emp.id if emp else None
            acq_date = None
            if request.form.get('acquisition_date'):
                try:
                    acq_date = datetime.strptime(request.form.get('acquisition_date'), '%Y-%m-%d').date()
                except:
                    pass
            exp_date = None
            if request.form.get('expiry_date'):
                try:
                    exp_date = datetime.strptime(request.form.get('expiry_date'), '%Y-%m-%d').date()
                except:
                    pass
            old_assigned_id = asset.assigned_employee_id
            asset.asset_name = request.form['asset_name']
            asset.serial_number = request.form['serial_number']
            asset.asset_tag = request.form.get('asset_tag') or None
            asset.category = request.form['category']
            asset.location = request.form['location']
            asset.site = request.form.get('site') or None
            asset.status = request.form['status']
            asset.barcode = request.form.get('barcode') or None
            asset.service_tag = request.form.get('service_tag') or None
            asset.model = request.form.get('model') or None
            asset.manufacturer = request.form.get('manufacturer') or None
            asset.business_impact = request.form.get('business_impact') or None
            asset.department = request.form.get('department') or None
            asset.os = request.form.get('os') or None
            asset.service_pack = request.form.get('service_pack') or None
            asset.ram_gb = request.form.get('ram_gb') or None
            asset.virtual_memory_gb = request.form.get('virtual_memory_gb') or None
            asset.processor_info = request.form.get('processor_info') or None
            asset.processor_manufacturer = request.form.get('processor_manufacturer') or None
            asset.cpu_speed_ghz = request.form.get('cpu_speed_ghz') or None
            asset.number_of_cores = int(request.form.get('number_of_cores')) if request.form.get('number_of_cores') else None
            asset.hdd_model = request.form.get('hdd_model') or None
            asset.hdd_serial = request.form.get('hdd_serial') or None
            asset.hdd_capacity_gb = request.form.get('hdd_capacity_gb') or None
            asset.ip_address = request.form.get('ip_address') or None
            asset.mac_address = request.form.get('mac_address') or None
            asset.nic = request.form.get('nic') or None
            asset.default_gateway = request.form.get('default_gateway') or None
            asset.network = request.form.get('network') or None
            asset.subnet_mask = request.form.get('subnet_mask') or None
            asset.dhcp_enabled = bool(int(request.form.get('dhcp_enabled', 0)))
            asset.dhcp_server = request.form.get('dhcp_server') or None
            asset.vendor_name = request.form.get('vendor_name') or None
            asset.purchase_cost = float(request.form.get('purchase_cost') or 0)
            asset.acquisition_date = acq_date
            asset.expiry_date = exp_date
            asset.warranty_expiry = request.form.get('warranty_expiry') or None
            asset.monitor = request.form.get('monitor') or None
            asset.comments = request.form.get('comments') or None
            asset.assigned_employee_id = assigned_id
            asset.assigned_to = assigned_name
            asset.updated_at = datetime.utcnow()
            db.session.commit()
            if assigned_id:
                assignment = AssetAssignment(
                    asset_id=asset.id,
                    employee_id=assigned_id,
                    notes="Asset re-assigned"
                )
                db.session.add(assignment)
                db.session.commit()
                # NEW: Notify only if assignment changed
                if assigned_id != old_assigned_id:
                    create_notification(
                        employee_id=assigned_id,
                        message=f"Asset '{asset.asset_name}' ({asset.serial_number}) has been assigned to you.",
                        link="/employee-dashboard"
                    )
                    db.session.commit()
            if 'document' in request.files:
                file = request.files['document']
                if file and file.filename != '' and allowed_file(file.filename):
                    filename = secure_filename(file.filename)
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    filename = f"{asset.id}_{timestamp}_{filename}"
                    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                    file.save(filepath)
                    document = AssetDocument(
                        asset_id=asset.id,
                        document_name=request.form.get('document_name', 'Asset Document'),
                        document_type=request.form.get('document_type', 'Other'),
                        file_path=f'uploads/documents/{filename}',
                        remarks=request.form.get('document_remarks'),
                        uploaded_by=session.get('employee_id')
                    )
                    db.session.add(document)
                    db.session.commit()
                    flash('Asset and document updated successfully!', 'success')
                else:
                    if file and file.filename != '':
                        flash('Asset updated! (Document skipped - invalid file type)', 'warning')
                    else:
                        flash('Asset updated successfully!', 'success')
            else:
                flash('Asset updated successfully!', 'success')
            generate_asset_qr(asset)
            log_action(action="UPDATE", entity_type="Asset", entity_id=asset.id,
                       details=f"Asset updated: {asset.asset_name} ({asset.serial_number})")
            db.session.commit()
            return redirect(url_for('list_assets'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating asset: {str(e)}', 'danger')
    return render_template('edit_asset.html', asset=asset, employees=employees)



@app.route('/delete-asset/<int:asset_id>', methods=['POST'])
@login_required
@admin_required
def delete_asset(asset_id):
    asset = Asset.query.get_or_404(asset_id)
    log_action(action="DELETE", entity_type="Asset", entity_id=asset.id,
               details=f"Asset deleted: {asset.asset_name} ({asset.serial_number})")
    db.session.delete(asset)
    db.session.commit()
    flash('Asset deleted.', 'danger')
    return redirect(url_for('list_assets'))



@app.route('/employees')
@login_required
@admin_required
def list_employees():
    employees = Employee.query.order_by(Employee.name).all()
    enriched = []
    for e in employees:
        assets_count = Asset.query.filter_by(assigned_employee_id=e.id).count()
        enriched.append({
            "id": e.id,
            "employee_id": e.employee_id,
            "name": e.name,
            "email": e.email or '-',
            "mobile": e.mobile or '-',
            "department": e.department,
            "designation": e.designation or '-',
            "assets_assigned": assets_count
        })
    return render_template('employees.html', employees=enriched)



@app.route('/add-employee', methods=['GET', 'POST'])
@login_required
@admin_required
def add_employee():
    if request.method == 'POST':
        emp = Employee(
            employee_id=request.form['employee_id'],
            name=request.form['name'],
            email=request.form.get('email') or None,
            mobile=request.form.get('mobile') or None,
            department=request.form['department'],
            designation=request.form.get('designation') or None,
            employee_type=request.form.get('employee_type', 'staff')
        )
        db.session.add(emp)
        db.session.commit()
        username = request.form.get('username') or emp.employee_id
        password = request.form.get('password', 'password123')
        user = User(username=username, password=password, role='employee', employee_id=emp.id)
        db.session.add(user)
        db.session.commit()
        flash(f'Employee added! Login: {username} / {password}', 'success')
        return redirect(url_for('list_employees'))
    return render_template('add_employee.html')



@app.route('/edit-employee/<int:emp_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_employee(emp_id):
    emp = Employee.query.get_or_404(emp_id)
    if request.method == 'POST':
        emp.employee_id = request.form['employee_id']
        emp.name = request.form['name']
        emp.email = request.form.get('email') or None
        emp.mobile = request.form.get('mobile') or None
        emp.department = request.form['department']
        emp.designation = request.form.get('designation') or None
        emp.employee_type = request.form.get('employee_type', 'staff')  # ✅ ADD THIS

        db.session.commit()
        flash('Employee updated!', 'success')
        return redirect(url_for('list_employees'))
    return render_template('edit_employee.html', employee=emp)



@app.route('/delete-employee/<int:emp_id>', methods=['POST'])
@login_required
@admin_required
def delete_employee(emp_id):
    emp = Employee.query.get_or_404(emp_id)
    db.session.delete(emp)
    db.session.commit()
    flash('Employee deleted.', 'danger')
    return redirect(url_for('list_employees'))



@app.route('/tracking')
@login_required
@admin_required
def tracking():
    search = request.args.get('search', '').strip()
    status = request.args.get('status', '')

    query = (
        db.session.query(AssetAssignment, Asset, Employee)
        .join(Asset, Asset.id == AssetAssignment.asset_id)
        .join(Employee, Employee.id == AssetAssignment.employee_id)
    )

    if search:
        like = f"%{search}%"
        query = query.filter(
            db.or_(
                Asset.asset_name.ilike(like),
                Asset.serial_number.ilike(like),
                Employee.name.ilike(like)
            )
        )
    if status == 'active':
        query = query.filter(AssetAssignment.returned_date == None)
    elif status == 'returned':
        query = query.filter(AssetAssignment.returned_date != None)

    assignments = query.order_by(AssetAssignment.assigned_date.desc()).all()
    return render_template('tracking.html', assignments=assignments)


# ===================== REPORTS HELPERS =====================

def build_report_data(report_type, category, department, status_filter, date_from, date_to):
    rows = []
    headers = []
    report_title = ''
    total_value = 0.0

    if report_type == 'inventory_summary':
        report_title = 'Inventory Summary by Category & Status'
        headers = ['Category', 'In Store', 'Deployed', 'In Maintenance', 'Total', 'Value (₹)']
        q = db.session.query(
            Asset.category,
            db.func.count(Asset.id).label('total'),
            db.func.sum(Asset.purchase_cost).label('value'),
            db.func.count(db.case((Asset.status == 'In Store', Asset.id))).label('in_store'),
            db.func.count(db.case((Asset.status == 'Deployed', Asset.id))).label('deployed'),
            db.func.count(db.case((Asset.status == 'In Maintenance', Asset.id))).label('maintenance')
        ).group_by(Asset.category)
        if category:
            q = q.filter(Asset.category == category)
        for r in q.order_by(db.desc('total')).all():
            val = r.value or 0
            total_value += val
            rows.append({
                'Category': r.category or '-',
                'In Store': r.in_store,
                'Deployed': r.deployed,
                'In Maintenance': r.maintenance,
                'Total': r.total,
                'Value (₹)': round(val, 2)
            })

    elif report_type == 'department_assets':
        report_title = 'Department-wise Asset Distribution'
        headers = ['Department', 'Total Assets', 'Total Value (₹)']
        q = db.session.query(
            Employee.department,
            db.func.count(Asset.id).label('assets'),
            db.func.sum(Asset.purchase_cost).label('value')
        ).outerjoin(Asset, Asset.assigned_employee_id == Employee.id) \
         .group_by(Employee.department)
        if department:
            q = q.filter(Employee.department == department)
        for r in q.order_by(db.desc('assets')).all():
            val = r.value or 0
            total_value += val
            rows.append({
                'Department': r.department or 'Unassigned',
                'Total Assets': r.assets,
                'Total Value (₹)': round(val, 2)
            })

    elif report_type == 'employee_custody':
        report_title = 'Employee Asset Custody'
        headers = ['Employee ID', 'Name', 'Department', 'Asset', 'Category', 'Status', 'Value (₹)']
        q = db.session.query(
            Employee.employee_id,
            Employee.name,
            Employee.department,
            Asset.asset_name,
            Asset.category,
            Asset.status,
            Asset.purchase_cost
        ).outerjoin(Asset, Asset.assigned_employee_id == Employee.id) \
         .order_by(Employee.name, Asset.asset_name)
        for r in q.limit(200).all():
            val = r.purchase_cost or 0
            total_value += val
            rows.append({
                'Employee ID': r.employee_id,
                'Name': r.name,
                'Department': r.department,
                'Asset': r.asset_name or '-',
                'Category': r.category or '-',
                'Status': r.status or '-',
                'Value (₹)': round(val, 2)
            })

    elif report_type == 'warranty_expiry':
        report_title = 'Warranty Expiry (Next 90 Days)'
        headers = ['Asset', 'Serial Number', 'Assigned To', 'Warranty Expiry', 'Value (₹)']
        cutoff_str = (datetime.now() + timedelta(days=90)).strftime('%Y-%m-%d')
        assets = Asset.query.filter(
            Asset.warranty_expiry != None,
            Asset.warranty_expiry != '',
            Asset.warranty_expiry <= cutoff_str
        ).order_by(Asset.warranty_expiry).limit(100).all()
        for a in assets:
            val = a.purchase_cost or 0
            total_value += val
            rows.append({
                'Asset': a.asset_name,
                'Serial Number': a.serial_number,
                'Assigned To': a.assigned_to or 'In Store',
                'Warranty Expiry': a.warranty_expiry,
                'Value (₹)': round(val, 2)
            })

    elif report_type == 'maintenance_backlog':
        report_title = 'Maintenance Backlog'
        headers = ['Asset', 'Serial Number', 'Category', 'Location', 'Last Updated', 'Value (₹)']
        q = Asset.query.filter(Asset.status == 'In Maintenance')
        if category:
            q = q.filter(Asset.category == category)
        for a in q.order_by(Asset.updated_at.asc()).limit(100).all():
            val = a.purchase_cost or 0
            total_value += val
            rows.append({
                'Asset': a.asset_name,
                'Serial Number': a.serial_number,
                'Category': a.category,
                'Location': a.location,
                'Last Updated': a.updated_at.strftime('%Y-%m-%d') if a.updated_at else '-',
                'Value (₹)': round(val, 2)
            })

    elif report_type == 'asset_reports':
        report_title = 'Employee Asset Issue Reports'
        headers = ['Date', 'Employee', 'Asset', 'Serial Number', 'Type', 'Message', 'Status']
        q = db.session.query(AssetReport, Asset, Employee) \
            .join(Asset, Asset.id == AssetReport.asset_id) \
            .join(Employee, Employee.id == AssetReport.employee_id) \
            .order_by(AssetReport.created_at.desc())
        if status_filter:
            q = q.filter(AssetReport.status == status_filter)
        for rep, ast, emp in q.limit(200).all():
            rows.append({
                'Date': rep.created_at.strftime('%Y-%m-%d %H:%M'),
                'Employee': emp.name,
                'Asset': ast.asset_name,
                'Serial Number': ast.serial_number,
                'Type': rep.report_type or '-',
                'Message': rep.message,
                'Status': rep.status
            })

    elif report_type == 'ticket_summary':
        report_title = 'Ticket Summary by Priority & Status'
        headers = ['Priority', 'Status', 'Count']
        q = db.session.query(
            Ticket.priority,
            Ticket.status,
            db.func.count(Ticket.id).label('count')
        ).group_by(Ticket.priority, Ticket.status) \
         .order_by(Ticket.priority, Ticket.status)
        for r in q.all():
            rows.append({
                'Priority': r.priority,
                'Status': r.status,
                'Count': r.count
            })

    return {
        'rows': rows,
        'headers': headers,
        'report_title': report_title,
        'total_value': total_value
    }


def export_report_excel(report_type, report_data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = report_type.replace('_', ' ').title()[:31]

    header_fill = PatternFill("solid", fgColor="003082")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal='center')

    headers = report_data['headers']
    rows = report_data['rows']

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    for row_idx, row_dict in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row_dict.get(header, ''))

    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=8) + 4
        ws.column_dimensions[col[0].column_letter].width = min(max_len, 50)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"AAI_IMS_{report_type}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/reports')
@login_required
@admin_required
def reports():
    report_type   = request.args.get('report_type', 'inventory_summary')
    category      = request.args.get('category', '')
    department    = request.args.get('department', '')
    status_filter = request.args.get('status_filter', '')
    date_from     = request.args.get('date_from', '')
    date_to       = request.args.get('date_to', '')
    export        = request.args.get('export', '')

    categories  = [c[0] for c in db.session.query(distinct(Asset.category)).order_by(Asset.category).all() if c[0]]
    departments = [d[0] for d in db.session.query(distinct(Employee.department)).order_by(Employee.department).all() if d[0]]

    report_data = build_report_data(report_type, category, department, status_filter, date_from, date_to)

    if export == 'excel':
        return export_report_excel(report_type, report_data)

    extra_params = '&'.join(filter(None, [
        f'category={category}'           if category      else '',
        f'department={department}'       if department    else '',
        f'status_filter={status_filter}' if status_filter else '',
        f'date_from={date_from}'         if date_from     else '',
        f'date_to={date_to}'             if date_to       else '',
    ]))

    return render_template(
        'reports.html',
        report_type=report_type,
        report_data=report_data,
        categories=categories,
        departments=departments,
        category=category,
        department=department,
        status_filter=status_filter,
        date_from=date_from,
        date_to=date_to,
        extra_params=extra_params
    )

@app.route('/resolve-report/<int:report_id>', methods=['POST'])
@login_required
@admin_required
def resolve_report(report_id):
    report = AssetReport.query.get_or_404(report_id)
    report.status = 'resolved'
    db.session.commit()
    flash('Report marked as resolved!', 'success')
    return redirect(url_for('reports') + '?report_type=asset_reports')

@app.route('/audit-logs')
@login_required
@admin_required
def audit_logs():
    logs = AuditLog.query.order_by(AuditLog.timestamp.desc()).limit(200).all()
    return render_template('audit_logs.html', logs=logs)



# ===================== TICKET ROUTES (ADMIN) =====================

# ===================== TICKET ROUTES (ADMIN) =====================


@app.route('/tickets')
@login_required
@admin_required
def tickets():
    all_tickets = Ticket.query.order_by(Ticket.raised_at.desc()).all()
    for ticket in all_tickets:
        if ticket.is_overdue() and not ticket.reminder_sent:
            print(f"🔔 REMINDER: Ticket {ticket.ticket_number} is overdue!")
            ticket.reminder_sent = True
            ticket.reminder_sent_at = datetime.utcnow()
            if ticket.assigned_to and ticket.assigned_to.email:
                send_email(
                    to=ticket.assigned_to.email,
                    subject=f"⚠️ Overdue Ticket: {ticket.ticket_number}",
                    html_body=f"""
                    <h3>Overdue Ticket Alert</h3>
                    <p>Hi <b>{ticket.assigned_to.name}</b>,</p>
                    <p>Ticket <b>{ticket.ticket_number}</b> - <b>{ticket.title}</b> is now overdue.</p>
                    <p>Priority: <b>{ticket.priority}</b> | Status: <b>{ticket.status}</b></p>
                    <p>Raised at: {ticket.raised_at.strftime('%d %b %Y, %I:%M %p')}</p>
                    <p>Please resolve this ticket as soon as possible.</p>
                    <br><p>- AAI Inventory Management System</p>
                    """
                )
    db.session.commit()
    all_employees = Employee.query.filter_by(employee_type='it_support').all()
    employee_stats = []
    for emp in all_employees:
        active_tickets = Ticket.query.filter(
            Ticket.assigned_to_employee_id == emp.id,
            db.or_(Ticket.status == 'Open', Ticket.status == 'In Progress')
        ).count()
        emp.current_workload = active_tickets
        workload_percent = (active_tickets / emp.max_tickets * 100) if emp.max_tickets > 0 else 0
        if not emp.is_available:
            status = 'unavailable'
        elif active_tickets >= emp.max_tickets:
            status = 'full'
        elif active_tickets >= emp.max_tickets * 0.7:
            status = 'busy'
        else:
            status = 'available'
        employee_stats.append({
            'id': emp.id,
            'name': emp.name,
            'department': emp.department,
            'active_tickets': active_tickets,
            'max_tickets': emp.max_tickets,
            'is_available': emp.is_available,
            'workload_percent': workload_percent,
            'status': status
        })
    db.session.commit()
    employee_stats.sort(key=lambda x: (x['status'] == 'unavailable', x['status'] == 'full', x['workload_percent']))
    stats = {
        'total': Ticket.query.count(),
        'open': Ticket.query.filter_by(status='Open').count(),
        'in_progress': Ticket.query.filter_by(status='In Progress').count(),
        'overdue': len([t for t in all_tickets if t.is_overdue()]),
        'unassigned': Ticket.query.filter_by(assigned_to_employee_id=None).count()
    }
    return render_template('tickets.html',
                         tickets=all_tickets,
                         stats=stats,
                         employee_stats=employee_stats,
                         all_employees=all_employees)


@app.route('/ticket/<int:id>')
@login_required
def view_ticket(id):
    ticket = Ticket.query.get_or_404(id)
    employees = Employee.query.filter_by(employee_type='it_support').all()
    return render_template('view_ticket.html', ticket=ticket, employees=employees)


@app.route('/debug-tickets')
@login_required
def debug_tickets():
    all_tickets = Ticket.query.all()
    ticket_data = []
    for ticket in all_tickets:
        ticket_data.append({
            'id': ticket.id,
            'ticket_number': ticket.ticket_number,
            'title': ticket.title,
            'status': ticket.status,
            'raised_by_employee_id': ticket.raised_by_employee_id,
            'raised_at': str(ticket.raised_at)
        })
    return jsonify({'total_tickets': len(all_tickets), 'tickets': ticket_data})


@app.route('/ticket/<int:id>/update', methods=['POST'])
@login_required
@admin_required
def update_ticket(id):
    ticket = Ticket.query.get_or_404(id)
    old_status = ticket.status
    old_assigned = ticket.assigned_to_employee_id

    ticket.status = request.form['status']
    assigned_id = request.form.get('assigned_to_employee_id')
    ticket.assigned_to_employee_id = int(assigned_id) if assigned_id else None

    if request.form['status'] == 'In Progress' and not ticket.assigned_at:
        ticket.assigned_at = datetime.utcnow()

    if request.form['status'] == 'Resolved' and not ticket.resolved_at:
        ticket.resolved_at = datetime.utcnow()
        ticket.resolution_notes = request.form.get('resolution_notes')

    if old_assigned != ticket.assigned_to_employee_id:
        assigned_to_name = ticket.assigned_to.name if ticket.assigned_to else 'Unassigned'
        log_action(action="ASSIGN", entity_type="Ticket", entity_id=ticket.id,
                   details=f"Ticket {ticket.ticket_number} assigned to {assigned_to_name}")

    if old_status != ticket.status:
        log_action(action="UPDATE", entity_type="Ticket", entity_id=ticket.id,
                   details=f"Ticket {ticket.ticket_number} status changed from {old_status} to {ticket.status}")

    db.session.commit()

    # Notify employee when ticket is assigned to them
    if ticket.assigned_to_employee_id and ticket.assigned_to_employee_id != old_assigned:
        create_notification(
            employee_id=ticket.assigned_to_employee_id,
            message=f"Ticket {ticket.ticket_number} - '{ticket.title}' has been assigned to you. Priority: {ticket.priority}",
            link=f"/ticket/{ticket.id}"
        )
        if ticket.assigned_to and ticket.assigned_to.email:
            send_email(
                to=ticket.assigned_to.email,
                subject=f"Ticket Assigned: {ticket.ticket_number}",
                html_body=f"""
                <h3>New Ticket Assigned</h3>
                <p>Hi <b>{ticket.assigned_to.name}</b>,</p>
                <p>A ticket has been assigned to you:</p>
                <ul>
                    <li><b>Ticket #:</b> {ticket.ticket_number}</li>
                    <li><b>Title:</b> {ticket.title}</li>
                    <li><b>Priority:</b> {ticket.priority}</li>
                    <li><b>Category:</b> {ticket.category or 'General'}</li>
                </ul>
                <p>Please login to the AAI IMS to view and action this ticket.</p>
                <br><p>- AAI Inventory Management System</p>
                """
            )
        db.session.commit()

    # Notify employee when their ticket is resolved
    if old_status != 'Resolved' and ticket.status == 'Resolved':
        create_notification(
            employee_id=ticket.raised_by_employee_id,
            message=f"Your ticket {ticket.ticket_number} - '{ticket.title}' has been resolved.",
            link=f"/ticket/{ticket.id}"
        )
        if ticket.raised_by and ticket.raised_by.email:
            send_email(
                to=ticket.raised_by.email,
                subject=f"Ticket Resolved: {ticket.ticket_number}",
                html_body=f"""
                <h3>Your Ticket Has Been Resolved</h3>
                <p>Hi <b>{ticket.raised_by.name}</b>,</p>
                <p>Your ticket <b>{ticket.ticket_number}</b> - <b>{ticket.title}</b> has been resolved.</p>
                <p><b>Resolution Notes:</b> {ticket.resolution_notes or 'No notes provided.'}</p>
                <br><p>- AAI Inventory Management System</p>
                """
            )
        db.session.commit()

    flash('Ticket updated successfully!', 'success')
    return redirect(f'/ticket/{id}')


# ===================== EMPLOYEE ROUTES =====================


@app.route('/employee-dashboard')
@login_required
def employee_dashboard():
    if session.get('role') != 'employee':
        return redirect(url_for('index'))
    emp_id = session.get('employee_id')
    employee = Employee.query.get_or_404(emp_id)
    assigned_assets = Asset.query.filter_by(assigned_employee_id=emp_id).all()
    my_tickets = Ticket.query.filter_by(raised_by_employee_id=emp_id).order_by(Ticket.raised_at.desc()).all()
    ticket_stats = {
        'total': len(my_tickets),
        'open': len([t for t in my_tickets if t.status == 'Open']),
        'in_progress': len([t for t in my_tickets if t.status == 'In Progress']),
        'resolved': len([t for t in my_tickets if t.status == 'Resolved'])
    }
    return render_template('employee_dashboard.html',
                         employee=employee,
                         assigned_assets=assigned_assets,
                         my_tickets=my_tickets,
                         ticket_stats=ticket_stats)


@app.route('/employee/raise-ticket', methods=['POST'])
@login_required
def employee_raise_ticket():
    new_ticket = Ticket(
        ticket_number=generate_ticket_number(),
        title=request.form['title'],
        description=request.form['description'],
        priority=request.form['priority'],
        category=request.form['category'],
        asset_id=request.form.get('asset_id') or None,
        raised_by_employee_id=request.form['raised_by_employee_id']
    )
    db.session.add(new_ticket)
    db.session.commit()

    # ✅ NEW: Notify all admins about the new ticket
    admin_users = User.query.filter_by(role='admin').all()
    for admin_user in admin_users:
        if admin_user.employee_id:
            create_notification(
                employee_id=admin_user.employee_id,
                message=f"🎫 New ticket {new_ticket.ticket_number} raised by {new_ticket.raised_by.name}: '{new_ticket.title}' [{new_ticket.priority}]",
                link=f"/ticket/{new_ticket.id}"
            )

    # ✅ NEW: Confirm to the employee who raised it
    create_notification(
        employee_id=new_ticket.raised_by_employee_id,
        message=f"✅ Your ticket {new_ticket.ticket_number} has been submitted. We'll get back to you soon.",
        link=f"/ticket/{new_ticket.id}"
    )
    db.session.commit()

    # Real-time socket event for admin dashboard
    socketio.emit('new_ticket', {
        'ticket_number': new_ticket.ticket_number,
        'title': new_ticket.title,
        'priority': new_ticket.priority,
        'raised_by': new_ticket.raised_by.name if new_ticket.raised_by else 'Unknown'
    })

    flash(f'Ticket {new_ticket.ticket_number} raised successfully!', 'success')
    return redirect('/employee-dashboard')


# ===================== EMPLOYEE ROUTES =====================

# ===================== NOTIFICATION ROUTES =====================

@app.route('/notifications')
@login_required
def notifications():
    emp_id = session.get('employee_id')
    role = session.get('role')

    # Admin has no employee_id — show empty state instead of redirecting
    if not emp_id:
        return render_template('notifications.html', notifications=[])

    notifs = Notification.query.filter_by(employee_id=emp_id) \
        .order_by(Notification.created_at.desc()).limit(50).all()

    for n in notifs:
        n.is_read = True
    db.session.commit()

    return render_template('notifications.html', notifications=notifs)


@app.route('/notifications/mark-read/<int:notif_id>', methods=['POST'])
@login_required
def mark_notification_read(notif_id):
    notif = Notification.query.get_or_404(notif_id)
    notif.is_read = True
    db.session.commit()
    return jsonify({'success': True})


@app.route('/notifications/unread-count')
@login_required
def unread_count():
    emp_id = session.get('employee_id')
    count = Notification.query.filter_by(
        employee_id=emp_id, is_read=False
    ).count() if emp_id else 0
    return jsonify({'count': count})



# ===================== ANALYTICS ROUTE =====================

@app.route('/analytics')
@login_required
@admin_required
def analytics():
    # Asset distribution by category
    asset_cats = db.session.query(
        Asset.category,
        db.func.count(Asset.id).label('count')
    ).group_by(Asset.category).all()

    # Asset distribution by status
    asset_statuses = db.session.query(
        Asset.status,
        db.func.count(Asset.id).label('count')
    ).group_by(Asset.status).all()

    # Ticket trends (last 7 days)
    today = datetime.utcnow().date()
    ticket_trends = []
    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        count = Ticket.query.filter(
            db.func.date(Ticket.raised_at) == day
        ).count()
        ticket_trends.append({'date': day.strftime('%d %b'), 'count': count})

    # Employee workload
    emp_workload = db.session.query(
        Employee.name,
        db.func.count(Ticket.id).label('tickets')
    ).outerjoin(Ticket, db.and_(
        Ticket.assigned_to_employee_id == Employee.id,
        db.or_(Ticket.status == 'Open', Ticket.status == 'In Progress')
    )).group_by(Employee.id).all()

    # Ticket status breakdown
    status_data = db.session.query(
        Ticket.status,
        db.func.count(Ticket.id).label('count')
    ).group_by(Ticket.status).all()

    # Ticket priority breakdown
    priority_data = db.session.query(
        Ticket.priority,
        db.func.count(Ticket.id).label('count')
    ).group_by(Ticket.priority).all()

    # Summary stats
    summary = {
        'total_assets': Asset.query.count(),
        'total_employees': Employee.query.count(),
        'total_tickets': Ticket.query.count(),
        'open_tickets': Ticket.query.filter_by(status='Open').count(),
        'overdue_tickets': len([t for t in Ticket.query.all() if t.is_overdue()]),
        'total_asset_value': db.session.query(db.func.sum(Asset.purchase_cost)).scalar() or 0
    }

    return render_template('analytics.html',
        asset_cats=asset_cats,
        asset_statuses=asset_statuses,
        ticket_trends=ticket_trends,
        emp_workload=emp_workload,
        status_data=status_data,
        priority_data=priority_data,
        summary=summary
    )



# ===================== API ROUTES =====================

@app.route('/api/search')
@login_required
def api_search():
    query = request.args.get('q', '').strip()
    if len(query) < 2:
        return jsonify([])
    like_pattern = f"%{query}%"
    assets = Asset.query.filter(
        db.or_(
            Asset.asset_name.ilike(like_pattern),
            Asset.serial_number.ilike(like_pattern),
            Asset.category.ilike(like_pattern),
            Asset.location.ilike(like_pattern)
        )
    ).limit(10).all()
    results = [{
        'name': asset.asset_name,
        'serial_number': asset.serial_number,
        'category': asset.category,
        'location': asset.location
    } for asset in assets]
    return jsonify(results)


# NEW: API endpoint for real-time stats refresh
@app.route('/api/stats')
@login_required
@admin_required
def api_stats():
    all_tickets = Ticket.query.all()
    return jsonify({
        'total_assets': Asset.query.count(),
        'total_tickets': Ticket.query.count(),
        'open_tickets': Ticket.query.filter_by(status='Open').count(),
        'in_progress_tickets': Ticket.query.filter_by(status='In Progress').count(),
        'overdue_tickets': len([t for t in all_tickets if t.is_overdue()]),
        'unassigned_tickets': Ticket.query.filter_by(assigned_to_employee_id=None).count()
    })

# ===================== AAI-IMSam AI ASSISTANT =====================

def imsam_query(question):
    """
    Rule-based NLP engine that maps natural language questions
    to live database queries and returns human-readable answers.
    """
    q = question.lower().strip()
    response = []

    # ── ASSET COUNT QUERIES ──
    if any(w in q for w in ['how many', 'count', 'total']) and \
       any(w in q for w in ['asset', 'assets', 'item', 'items']):

        if any(w in q for w in ['deploy', 'deployed', 'issued']):
            count = Asset.query.filter_by(status='Deployed').count()
            val = db.session.query(db.func.sum(Asset.purchase_cost))\
                    .filter(Asset.status=='Deployed').scalar() or 0
            return f"📦 <b>{count}</b> assets are currently deployed with a total value of <b>₹{val:,.0f}</b>."

        elif any(w in q for w in ['maintenance', 'repair', 'broken']):
            count = Asset.query.filter_by(status='In Maintenance').count()
            return f"🔧 <b>{count}</b> assets are currently in maintenance."

        elif any(w in q for w in ['store', 'storage', 'available', 'unassigned']):
            count = Asset.query.filter_by(status='In Store').count()
            return f"🏪 <b>{count}</b> assets are currently in store (unassigned)."

        else:
            total = Asset.query.count()
            deployed = Asset.query.filter_by(status='Deployed').count()
            in_store = Asset.query.filter_by(status='In Store').count()
            maintenance = Asset.query.filter_by(status='In Maintenance').count()
            return (f"📊 Total assets: <b>{total}</b><br>"
                    f"&nbsp;&nbsp;• Deployed: <b>{deployed}</b><br>"
                    f"&nbsp;&nbsp;• In Store: <b>{in_store}</b><br>"
                    f"&nbsp;&nbsp;• In Maintenance: <b>{maintenance}</b>")

    # ── CATEGORY SPECIFIC ──
    categories = ['laptop', 'desktop', 'printer', 'switch', 'router',
                  'server', 'ups', 'monitor', 'network', 'computer']
    matched_cat = next((c for c in categories if c in q), None)

    if matched_cat:
        like = f"%{matched_cat}%"
        assets = Asset.query.filter(
            db.or_(Asset.category.ilike(like), Asset.asset_name.ilike(like))
        ).all()
        if assets:
            deployed = sum(1 for a in assets if a.status == 'Deployed')
            maintenance = sum(1 for a in assets if a.status == 'In Maintenance')
            in_store = sum(1 for a in assets if a.status == 'In Store')
            return (f"💻 Found <b>{len(assets)}</b> {matched_cat}(s):<br>"
                    f"&nbsp;&nbsp;• Deployed: <b>{deployed}</b><br>"
                    f"&nbsp;&nbsp;• In Store: <b>{in_store}</b><br>"
                    f"&nbsp;&nbsp;• In Maintenance: <b>{maintenance}</b>")
        else:
            return f"🔍 No assets found matching '<b>{matched_cat}</b>'."

    # ── WARRANTY QUERIES ──
    if any(w in q for w in ['warranty', 'expire', 'expiry', 'expiring']):
        days = 90
        if '30' in q or 'month' in q: days = 30
        elif '60' in q: days = 60
        cutoff = (datetime.now() + timedelta(days=days)).strftime('%Y-%m-%d')
        assets = Asset.query.filter(
            Asset.warranty_expiry != None,
            Asset.warranty_expiry != '',
            Asset.warranty_expiry <= cutoff
        ).order_by(Asset.warranty_expiry).limit(10).all()
        if assets:
            lines = [f"⚠️ <b>{len(assets)}</b> asset(s) warranty expiring within {days} days:<br>"]
            for a in assets:
                lines.append(f"&nbsp;&nbsp;• {a.asset_name} ({a.serial_number}) — <b>{a.warranty_expiry}</b><br>")
            return ''.join(lines)
        return f"✅ No assets with warranty expiring in the next {days} days."

    # ── DEPARTMENT QUERIES ──
    if any(w in q for w in ['department', 'dept', 'division', 'section']):
        results = db.session.query(
            Employee.department,
            db.func.count(Asset.id).label('count')
        ).outerjoin(Asset, Asset.assigned_employee_id == Employee.id)\
         .group_by(Employee.department)\
         .order_by(db.desc('count')).all()
        if results:
            lines = ["🏢 <b>Assets by Department:</b><br>"]
            for dept, count in results:
                lines.append(f"&nbsp;&nbsp;• {dept}: <b>{count}</b> assets<br>")
            return ''.join(lines)

    # ── EMPLOYEE / WHO HAS MOST ──
    if any(w in q for w in ['who', 'employee', 'person', 'staff', 'most assets']):
        results = db.session.query(
            Employee.name,
            Employee.department,
            db.func.count(Asset.id).label('count')
        ).outerjoin(Asset, Asset.assigned_employee_id == Employee.id)\
         .group_by(Employee.id)\
         .order_by(db.desc('count')).limit(5).all()
        if results:
            lines = ["👤 <b>Top asset holders:</b><br>"]
            for name, dept, count in results:
                lines.append(f"&nbsp;&nbsp;• {name} ({dept}): <b>{count}</b> assets<br>")
            return ''.join(lines)

    # ── TICKET QUERIES ──
    if any(w in q for w in ['ticket', 'tickets', 'issue', 'issues']):
        total = Ticket.query.count()
        open_t = Ticket.query.filter_by(status='Open').count()
        inprog = Ticket.query.filter_by(status='In Progress').count()
        resolved = Ticket.query.filter_by(status='Resolved').count()
        all_t = Ticket.query.all()
        overdue = sum(1 for t in all_t if t.is_overdue())
        return (f"🎫 <b>Ticket Summary:</b><br>"
                f"&nbsp;&nbsp;• Total: <b>{total}</b><br>"
                f"&nbsp;&nbsp;• Open: <b>{open_t}</b><br>"
                f"&nbsp;&nbsp;• In Progress: <b>{inprog}</b><br>"
                f"&nbsp;&nbsp;• Resolved: <b>{resolved}</b><br>"
                f"&nbsp;&nbsp;• Overdue: <b>{overdue}</b>")

    # ── ASSET VALUE ──
    if any(w in q for w in ['value', 'cost', 'worth', 'price', 'expensive']):
        total_val = db.session.query(db.func.sum(Asset.purchase_cost)).scalar() or 0
        deployed_val = db.session.query(db.func.sum(Asset.purchase_cost))\
                         .filter(Asset.status=='Deployed').scalar() or 0
        top = Asset.query.order_by(Asset.purchase_cost.desc()).first()
        lines = [f"💰 <b>Asset Portfolio Value:</b><br>",
                 f"&nbsp;&nbsp;• Total: <b>₹{total_val:,.0f}</b><br>",
                 f"&nbsp;&nbsp;• Deployed Value: <b>₹{deployed_val:,.0f}</b><br>"]
        if top:
            lines.append(f"&nbsp;&nbsp;• Most expensive: <b>{top.asset_name}</b> (₹{top.purchase_cost:,.0f})")
        return ''.join(lines)

    # ── MAINTENANCE BACKLOG ──
    if any(w in q for w in ['maintenance', 'repair', 'broken', 'backlog']):
        assets = Asset.query.filter_by(status='In Maintenance')\
                     .order_by(Asset.updated_at.asc()).all()
        if assets:
            lines = [f"🔧 <b>{len(assets)}</b> asset(s) in maintenance:<br>"]
            for a in assets[:5]:
                lines.append(f"&nbsp;&nbsp;• {a.asset_name} ({a.serial_number}) — {a.location}<br>")
            if len(assets) > 5:
                lines.append(f"&nbsp;&nbsp;<i>...and {len(assets)-5} more</i>")
            return ''.join(lines)
        return "✅ No assets currently in maintenance."

    # ── HELP / GREETING ──
    if any(w in q for w in ['hi', 'hello', 'hey', 'help', 'what can you']):
        return (f"👋 Hi! I'm <b>AAI-IMSam</b>, your Smart Asset Manager.<br><br>"
                f"You can ask me things like:<br>"
                f"&nbsp;&nbsp;• 'How many assets are deployed?'<br>"
                f"&nbsp;&nbsp;• 'Show warranty expiring soon'<br>"
                f"&nbsp;&nbsp;• 'How many laptops do we have?'<br>"
                f"&nbsp;&nbsp;• 'Which department has the most assets?'<br>"
                f"&nbsp;&nbsp;• 'Show ticket summary'<br>"
                f"&nbsp;&nbsp;• 'What is the total asset value?'")
    # ── EMPLOYEE-SPECIFIC ASSET QUERY ──
    # "assets assigned to tony stark" / "what does john have"
    if any(w in q for w in ['assigned to', 'belongs to', 'what does', 'assets of', 'has how many']):
        # Extract employee name — everything after 'to', 'of', 'does'
        name_part = ''
        for trigger in ['assigned to ', 'belongs to ', 'what does ', 'assets of ']:
            if trigger in q:
                name_part = q.split(trigger)[-1].replace('have', '').replace('has', '').strip()
                break
        if name_part:
            emp = Employee.query.filter(Employee.name.ilike(f'%{name_part}%')).first()
            if emp:
                assets = Asset.query.filter_by(assigned_employee_id=emp.id).all()
                if assets:
                    lines = [f"👤 <b>{emp.name}</b> ({emp.department}) has <b>{len(assets)}</b> asset(s):<br>"]
                    for a in assets:
                        lines.append(f"&nbsp;&nbsp;• {a.asset_name} ({a.serial_number}) — "
                                     f"<span style='color:#f59e0b'>{a.status}</span><br>")
                    return ''.join(lines)
                else:
                    return f"👤 <b>{emp.name}</b> currently has no assets assigned."
            else:
                return f"🔍 No employee found matching '<b>{name_part}</b>'."

    # ── SPECIFIC ASSET LOOKUP ──
    # "find asset ASSET-IT-001" / "show serial #001"
    if any(w in q for w in ['find asset', 'search asset', 'show asset', 'serial', 'asset id']):
        # Try to extract serial or asset name
        words = q.split()
        for word in words:
            if len(word) > 3:
                asset = Asset.query.filter(
                    db.or_(
                        Asset.serial_number.ilike(f'%{word}%'),
                        Asset.asset_name.ilike(f'%{word}%')
                    )
                ).first()
                if asset:
                    owner = Employee.query.get(asset.assigned_employee_id) if asset.assigned_employee_id else None
                    return (f"🔍 <b>{asset.asset_name}</b><br>"
                            f"&nbsp;&nbsp;• Serial: <b>{asset.serial_number}</b><br>"
                            f"&nbsp;&nbsp;• Category: {asset.category}<br>"
                            f"&nbsp;&nbsp;• Status: <b>{asset.status}</b><br>"
                            f"&nbsp;&nbsp;• Location: {asset.location}<br>"
                            f"&nbsp;&nbsp;• Assigned to: <b>{'None' if not owner else owner.name}</b><br>"
                            f"&nbsp;&nbsp;• Warranty: {asset.warranty_expiry or 'N/A'}")

    # ── LOCATION QUERIES ──
    # "assets in 3rd floor" / "what is in operations room"
    if any(w in q for w in ['location', 'floor', 'room', 'where', 'placed']):
        loc_part = ''
        for trigger in ['in ', 'at ', 'on ']:
            if trigger in q:
                loc_part = q.split(trigger)[-1].strip()
                break
        if loc_part:
            assets = Asset.query.filter(Asset.location.ilike(f'%{loc_part}%')).all()
            if assets:
                lines = [f"📍 <b>{len(assets)}</b> asset(s) found at '<b>{loc_part}</b>':<br>"]
                for a in assets[:8]:
                    lines.append(f"&nbsp;&nbsp;• {a.asset_name} ({a.serial_number}) — {a.status}<br>")
                if len(assets) > 8:
                    lines.append(f"&nbsp;&nbsp;<i>...and {len(assets)-8} more</i>")
                return ''.join(lines)
            return f"📍 No assets found at location matching '<b>{loc_part}</b>'."

    # ── AGE / OLD ASSETS ──
    # "assets older than 3 years" / "oldest assets"
    if any(w in q for w in ['old', 'older', 'age', 'oldest', 'years']):
        years = 3
        for word in q.split():
            if word.isdigit():
                years = int(word)
                break
        cutoff = (datetime.now() - timedelta(days=years*365)).strftime('%Y-%m-%d')
        assets = Asset.query.filter(
            Asset.purchase_date != None,
            Asset.purchase_date != '',
            Asset.purchase_date <= cutoff
        ).order_by(Asset.purchase_date.asc()).limit(8).all()
        if assets:
            lines = [f"📅 <b>{len(assets)}</b> asset(s) older than <b>{years} year(s)</b>:<br>"]
            for a in assets:
                lines.append(f"&nbsp;&nbsp;• {a.asset_name} ({a.serial_number}) — purchased {a.purchase_date}<br>")
            return ''.join(lines)
        return f"✅ No assets older than {years} year(s) found."

    # ── UNASSIGNED ASSETS ──
    # "show unassigned assets" / "what is available"
    if any(w in q for w in ['unassigned', 'available', 'free', 'no one', 'nobody']):
        assets = Asset.query.filter_by(status='In Store').all()
        if assets:
            lines = [f"🏪 <b>{len(assets)}</b> unassigned asset(s) in store:<br>"]
            for a in assets[:8]:
                lines.append(f"&nbsp;&nbsp;• {a.asset_name} ({a.serial_number}) — {a.category}<br>")
            if len(assets) > 8:
                lines.append(f"&nbsp;&nbsp;<i>...and {len(assets)-8} more</i>")
            return ''.join(lines)
        return "✅ No unassigned assets — everything is deployed."

    # ── CRITICAL / HIGH PRIORITY TICKETS ──
    # "show critical tickets" / "urgent issues"
    if any(w in q for w in ['critical', 'urgent', 'high priority', 'important']):
        tickets = Ticket.query.filter(
            Ticket.priority.in_(['Critical', 'High']),
            Ticket.status != 'Resolved'
        ).order_by(Ticket.raised_at.desc()).limit(5).all()
        if tickets:
            lines = [f"🚨 <b>{len(tickets)}</b> critical/high priority open ticket(s):<br>"]
            for t in tickets:
                lines.append(f"&nbsp;&nbsp;• [{t.ticket_number}] {t.title} — "
                             f"<b>{t.priority}</b> · {t.status}<br>")
            return ''.join(lines)
        return "✅ No critical or high priority open tickets."

    # ── EMPLOYEE LIST / INFO ──
    # "list all employees" / "how many employees"
    if any(w in q for w in ['employee', 'employees', 'staff', 'team', 'workers']):
        if any(w in q for w in ['how many', 'count', 'total', 'number']):
            count = Employee.query.count()
            depts = db.session.query(Employee.department).distinct().count()
            return f"👥 Total <b>{count}</b> employees across <b>{depts}</b> departments."
        else:
            emps = Employee.query.order_by(Employee.department).all()
            if emps:
                lines = [f"👥 <b>Employee Directory ({len(emps)}):</b><br>"]
                for e in emps:
                    asset_count = Asset.query.filter_by(assigned_employee_id=e.id).count()
                    lines.append(f"&nbsp;&nbsp;• {e.name} — {e.department} "
                                f"(<b>{asset_count}</b> assets)<br>")
                return ''.join(lines)

    # ── AUDIT / RECENT ACTIVITY ──
    # "recent activity" / "latest changes"
    if any(w in q for w in ['recent', 'latest', 'last', 'activity', 'audit', 'log', 'history']):
        try:
            logs = AuditLog.query.order_by(AuditLog.timestamp.desc()).limit(5).all()
            if logs:
                lines = ["🕐 <b>Recent Activity:</b><br>"]
                for log in logs:
                    lines.append(f"&nbsp;&nbsp;• {log.action} — "
                                f"<span style='color:#94a3b8'>{log.timestamp.strftime('%d %b %H:%M')}</span><br>")
                return ''.join(lines)
        except:
            return "📋 Audit log query not available."

    # ── FALLBACK ──
    return ("🤔 I didn't quite understand that. Try asking:<br>"
            "&nbsp;&nbsp;• 'How many assets are deployed?'<br>"
            "&nbsp;&nbsp;• 'Show warranty expiring soon'<br>"
            "&nbsp;&nbsp;• 'Which department has most assets?'")


@app.route('/imsam', methods=['POST'])
@login_required
def imsam():
    data = request.get_json()
    question = data.get('question', '').strip()
    if not question:
        return jsonify({'answer': "Please type a question."})
    try:
        answer = imsam_query(question)
    except Exception as e:
        answer = f"⚠️ Error processing query: {str(e)}"
    return jsonify({'answer': answer})

@app.route('/bulk-import', methods=['GET', 'POST'])
@login_required
@admin_required
def bulk_import():
    if request.method == 'POST':
        file = request.files.get('file')
        if not file or file.filename == '':
            flash('No file selected.', 'danger')
            return redirect('/bulk-import')

        try:
            wb = openpyxl.load_workbook(file)
            assets_added = 0
            employees_added = 0
            skipped = 0

            # ── EMPLOYEES SHEET ──────────────────────────────────────
            if 'Employees' in wb.sheetnames:
                ws = wb['Employees']
                headers = [str(c.value).strip().lower().replace(' ', '_') if c.value else '' for c in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue
                    r = dict(zip(headers, row))
                    emp_id = str(r.get('employee_id', '') or '').strip()
                    name   = str(r.get('name', '') or '').strip()
                    if not emp_id or not name:
                        skipped += 1
                        continue
                    if Employee.query.filter_by(employee_id=emp_id).first():
                        skipped += 1
                        continue
                    emp = Employee(
                        employee_id  = emp_id,
                        name         = name,
                        email        = str(r.get('email', '') or '').strip() or None,
                        mobile       = str(r.get('mobile', '') or '').strip() or None,
                        department   = str(r.get('department', '') or 'General').strip(),
                        designation  = str(r.get('designation', '') or '').strip() or None,
                    )
                    db.session.add(emp)
                    db.session.flush()  # get emp.id immediately
                    # create login user
                    uname = emp_id.lower()
                    if not User.query.filter_by(username=uname).first():
                        user = User(username=uname, password='password123',
                                    role='employee', employee_id=emp.id)
                        db.session.add(user)
                    employees_added += 1

            # ── ASSETS SHEET ─────────────────────────────────────────
            if 'Assets' in wb.sheetnames:
                ws = wb['Assets']
                headers = [str(c.value).strip().lower().replace(' ', '_') if c.value else '' for c in ws[1]]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue
                    r = dict(zip(headers, row))
                    asset_name    = str(r.get('asset_name', '') or '').strip()
                    serial_number = str(r.get('serial_number', '') or '').strip()
                    category      = str(r.get('category', '') or '').strip()
                    location      = str(r.get('location', '') or '').strip()

                    if not asset_name or not serial_number or not category or not location:
                        skipped += 1
                        continue
                    if Asset.query.filter_by(serial_number=serial_number).first():
                        skipped += 1
                        continue

                    # parse dates safely
                    def parse_date(val):
                        if not val:
                            return None
                        try:
                            if isinstance(val, datetime):
                                return val.date()
                            return datetime.strptime(str(val).strip(), '%Y-%m-%d').date()
                        except:
                            return None

                    assigned_name = str(r.get('assigned_to', '') or '').strip() or None
                    employee      = get_employee_by_name(assigned_name) if assigned_name else None

                    asset = Asset(
                        asset_name    = asset_name,
                        serial_number = serial_number,
                        asset_tag     = str(r.get('asset_tag', '') or '').strip() or None,
                        category      = category,
                        location      = location,
                        site          = str(r.get('site', '') or '').strip() or None,
                        status        = str(r.get('status', 'In Store') or 'In Store').strip(),
                        model         = str(r.get('model', '') or '').strip() or None,
                        manufacturer  = str(r.get('manufacturer', '') or '').strip() or None,
                        department    = str(r.get('department', '') or '').strip() or None,
                        vendor_name   = str(r.get('vendor_name', '') or '').strip() or None,
                        purchase_cost = float(r.get('purchase_cost') or 0),
                        acquisition_date = parse_date(r.get('acquisition_date')),
                        warranty_expiry  = str(r.get('warranty_expiry', '') or '').strip() or None,
                        assigned_employee_id = employee.id if employee else None,
                        assigned_to   = assigned_name,
                        comments      = str(r.get('comments', '') or '').strip() or None,
                    )
                    db.session.add(asset)
                    assets_added += 1

            db.session.commit()
            flash(f'✅ Import complete! {assets_added} assets and {employees_added} employees added. ({skipped} rows skipped)', 'success')

        except Exception as e:
            db.session.rollback()
            flash(f'❌ Import failed: {str(e)}', 'danger')

        return redirect('/assets')

    return render_template('bulk_import.html')

# ===================== IAMSAM CHATBOT ROUTE =====================

@app.route('/api/chat', methods=['POST'])
@login_required
def iamsam_chat():
    from datetime import datetime, timedelta
    import random as rnd

    data = request.get_json()
    user_msg = (data.get('message') or '').strip().lower()
    role = session.get('role', 'employee')

    if not user_msg:
        return jsonify({'reply': "Hey there! I'm Sam. Ask me anything about assets, tickets, or employees."})

    def fmt_currency(val):
        return f"₹{val:,.2f}"

    def get_asset_stats():
        total    = Asset.query.count()
        in_store = Asset.query.filter_by(status='In Store').count()
        deployed = Asset.query.filter_by(status='Deployed').count()
        maint    = Asset.query.filter_by(status='In Maintenance').count()
        total_val = db.session.query(db.func.sum(Asset.purchase_cost)).scalar() or 0
        return total, in_store, deployed, maint, total_val

    def get_ticket_stats():
        total    = Ticket.query.count()
        open_t   = Ticket.query.filter_by(status='Open').count()
        inprog   = Ticket.query.filter_by(status='In Progress').count()
        resolved = Ticket.query.filter_by(status='Resolved').count()
        closed   = Ticket.query.filter_by(status='Closed').count()
        overdue  = sum(1 for t in Ticket.query.filter(Ticket.status.notin_(['Resolved','Closed'])).all() if t.is_overdue())
        return total, open_t, inprog, resolved, closed, overdue

    def get_employee_stats():
        total = Employee.query.count()
        with_assets = db.session.query(db.func.count(db.func.distinct(Asset.assigned_employee_id))) \
                        .filter(Asset.assigned_employee_id != None).scalar() or 0
        return total, with_assets

    def get_expiring_warranties(days=30):
        cutoff = (datetime.now() + timedelta(days=days)).strftime('%Y-%m-%d')
        return Asset.query.filter(
            Asset.warranty_expiry != None,
            Asset.warranty_expiry != '',
            Asset.warranty_expiry <= cutoff
        ).order_by(Asset.warranty_expiry).limit(5).all()

    def match(keywords):
        return any(k in user_msg for k in keywords)

    reply = None

    # ══ 1. GREETINGS & IDENTITY ══
    if match(['hello','hi','hey','helo','namaste','good morning','good afternoon','good evening']):
        hour = datetime.now().hour
        greeting = 'Good morning' if hour < 12 else 'Good afternoon' if hour < 17 else 'Good evening'
        name = session.get('username', 'there')
        reply = f"{greeting}, {name}! 👋 I'm <b>Sam</b>, your AAI IMS assistant. I can help you with assets, tickets, employees, costs, and analytics. What would you like to know?"

    elif match(['who are you','what are you','your name','introduce yourself','what can you do']):
        reply = ("I'm <b>IAMSam</b> — the Inventory Asset Management assistant for AAI IMS! 🤖<br>"
                 "I can help you with:<br>• 📦 Asset counts, statuses & costs<br>"
                 "• 🎫 Ticket summaries & overdue alerts<br>• 👥 Employee & department info<br>"
                 "• ⚠️ Warranty expiry alerts<br>• 📊 Quick analytics<br>Just ask away!")

    elif match(['thank','thanks','thank you','धन्यवाद']):
        reply = "You're most welcome! 😊 Is there anything else I can help you with?"

    elif match(['bye','goodbye','see you','tata','cya']):
        reply = "Goodbye! 👋 If you need me again, I'm always right here. Have a great day!"

    elif match(['help','what can i ask','what do you know','commands']):
        reply = ("Here's what you can ask me about 📋<br><br>"
                 "<b>📦 Assets:</b> How many assets? / Total value / Deployed / In store / Maintenance<br>"
                 "<b>🎫 Tickets:</b> Open tickets / Overdue / Critical / How to raise<br>"
                 "<b>👥 Employees:</b> Count / Who has most assets / Departments<br>"
                 "<b>💰 Costs:</b> Total inventory value / Deployed value<br>"
                 "<b>⚠️ Alerts:</b> Warranties expiring / Overdue tickets<br>"
                 "<b>🔗 Navigation:</b> Where is Assets/Reports/Analytics?")

    # ══ 2. ASSET QUERIES ══
    elif match(['how many asset','total asset','asset count','number of asset','asset total']):
        total, in_store, deployed, maint, total_val = get_asset_stats()
        reply = (f"We currently have <b>{total} assets</b> in the system 📦<br>"
                 f"• In Store: <b>{in_store}</b><br>• Deployed: <b>{deployed}</b><br>"
                 f"• In Maintenance: <b>{maint}</b><br>• Total Value: <b>{fmt_currency(total_val)}</b>")

    elif match(['deployed asset','how many deployed','assets deployed']):
        _, _, deployed, _, _ = get_asset_stats()
        val = db.session.query(db.func.sum(Asset.purchase_cost)).filter_by(status='Deployed').scalar() or 0
        reply = f"There are currently <b>{deployed} deployed assets</b> out in the field 🚀<br>Deployed value: <b>{fmt_currency(val)}</b>"

    elif match(['in store','assets in store','stored asset','how many in store']):
        _, in_store, _, _, _ = get_asset_stats()
        val = db.session.query(db.func.sum(Asset.purchase_cost)).filter_by(status='In Store').scalar() or 0
        reply = f"We have <b>{in_store} assets currently in store</b> 🏬<br>In-store value: <b>{fmt_currency(val)}</b>"

    elif match(['maintenance','in maintenance','under repair','being repaired']):
        _, _, _, maint, _ = get_asset_stats()
        assets = Asset.query.filter_by(status='In Maintenance').limit(5).all()
        names = ', '.join(a.asset_name for a in assets) if assets else 'None'
        reply = f"<b>{maint} assets</b> are currently under maintenance 🔧<br>Recent: {names}"

    elif match(['unassigned asset','not assigned','no employee','free asset']):
        count = Asset.query.filter(Asset.assigned_employee_id == None, Asset.status == 'In Store').count()
        reply = f"There are <b>{count} unassigned assets</b> in store with no employee linked 📦<br>Head to <a href='/assets' style='color:#ff8a3d;'>Assets</a> to assign them."

    elif match(['asset value','total value','inventory worth','inventory value','how much worth']):
        _, _, _, _, total_val = get_asset_stats()
        dep_val   = db.session.query(db.func.sum(Asset.purchase_cost)).filter_by(status='Deployed').scalar() or 0
        store_val = db.session.query(db.func.sum(Asset.purchase_cost)).filter_by(status='In Store').scalar() or 0
        reply = (f"Here's the financial picture of our inventory 💰<br>"
                 f"• <b>Total Inventory Value:</b> {fmt_currency(total_val)}<br>"
                 f"• Deployed Assets Value: {fmt_currency(dep_val)}<br>"
                 f"• In-Store Assets Value: {fmt_currency(store_val)}")

    elif match(['category','categories','asset type','types of asset']):
        cats = db.session.query(Asset.category, db.func.count(Asset.id).label('cnt')) \
                 .group_by(Asset.category).order_by(db.desc('cnt')).limit(8).all()
        if cats:
            lines = '<br>'.join(f"• {c.category}: <b>{c.cnt}</b>" for c in cats)
            reply = f"Asset categories in your system 🗂️<br>{lines}"
        else:
            reply = "No asset categories found yet. Try adding some assets first!"

    # ══ 3. TICKET QUERIES ══
    elif match(['how many ticket','total ticket','ticket count','ticket summary']):
        total, open_t, inprog, resolved, closed, overdue = get_ticket_stats()
        reply = (f"Here's the current ticket snapshot 🎫<br>"
                 f"• Total: <b>{total}</b> | Open: <b>{open_t}</b><br>"
                 f"• In Progress: <b>{inprog}</b> | Resolved: <b>{resolved}</b><br>"
                 f"• ⚠️ Overdue (>2 hrs): <b>{overdue}</b>")

    elif match(['open ticket','pending ticket','unresolved ticket']):
        _, open_t, _, _, _, _ = get_ticket_stats()
        tickets = Ticket.query.filter_by(status='Open').order_by(Ticket.raised_at.desc()).limit(3).all()
        lines = '<br>'.join(f"• [{t.ticket_number}] {t.title} ({t.priority})" for t in tickets) if tickets else "None at the moment."
        reply = f"There are <b>{open_t} open tickets</b> right now 🔓<br><br>Latest:<br>{lines}"

    elif match(['overdue ticket','sla breach','delayed ticket','not resolved']):
        _, _, _, _, _, overdue = get_ticket_stats()
        if overdue == 0:
            reply = "Great news! ✅ No overdue tickets right now. Everything is within the 2-hour SLA."
        else:
            tickets = [t for t in Ticket.query.filter(Ticket.status.notin_(['Resolved','Closed'])).all() if t.is_overdue()][:3]
            lines = '<br>'.join(f"• [{t.ticket_number}] {t.title} — {t.time_since_raised()}" for t in tickets)
            reply = (f"⚠️ <b>{overdue} overdue tickets</b> are breaching the 2-hour SLA!<br><br>{lines}<br><br>"
                     f"<a href='/tickets' style='color:#ff8a3d;'>→ Manage Tickets</a>")

    elif match(['critical ticket','high priority ticket','urgent ticket']):
        count = Ticket.query.filter(Ticket.priority.in_(['High','Critical']), Ticket.status == 'Open').count()
        reply = f"There are <b>{count} high/critical priority open tickets</b> 🚨 <a href='/tickets' style='color:#ff8a3d;'>Review now →</a>"

    elif match(['raise ticket','create ticket','new ticket','log ticket','how to raise']):
        reply = "To raise a ticket 🎫, go to your <b>Dashboard</b> and click <b>'Raise Ticket'</b>. Fill in the asset, issue description, and priority — it'll be assigned to a technician automatically!"

    # ══ 4. EMPLOYEE QUERIES ══
    elif match(['how many employee','total employee','employee count','staff count']):
        total, with_assets = get_employee_stats()
        reply = (f"We have <b>{total} employees</b> registered 👥<br>"
                 f"• With assigned assets: <b>{with_assets}</b><br>"
                 f"• Without assets: <b>{total - with_assets}</b>")

    elif match(['most asset','who has most','highest asset','top employee','most assigned']):
        result = db.session.query(Employee.name, db.func.count(Asset.id).label('cnt')) \
                   .join(Asset, Asset.assigned_employee_id == Employee.id) \
                   .group_by(Employee.id).order_by(db.desc('cnt')).first()
        reply = f"The employee with the most assets is <b>{result.name}</b> with <b>{result.cnt} assets</b> 🏆" if result else "No assets have been assigned to any employee yet."

    elif match(['department','departments','which department']):
        depts = db.session.query(Employee.department, db.func.count(Employee.id).label('cnt')) \
                  .group_by(Employee.department).order_by(db.desc('cnt')).limit(6).all()
        lines = '<br>'.join(f"• {d.department}: <b>{d.cnt} employees</b>" for d in depts) if depts else "No data found."
        reply = f"Departments in AAI IMS 🏢<br>{lines}"

    # ══ 5. WARRANTY ALERTS ══
    elif match(['warranty','warranties','expir','expiring soon']):
        assets_30 = get_expiring_warranties(30)
        assets_90 = get_expiring_warranties(90)
        if not assets_90:
            reply = "✅ No warranties expiring in the next 90 days. You're all good!"
        else:
            lines = '<br>'.join(f"• <b>{a.asset_name}</b> (SN: {a.serial_number}) — expires {a.warranty_expiry}" for a in assets_30[:4])
            reply = (f"⚠️ <b>{len(assets_30)} warranties</b> expiring in 30 days, "
                     f"<b>{len(assets_90)}</b> in 90 days!<br><br>{lines}<br><br>"
                     f"<a href='/reports?report_type=warranty_expiry' style='color:#ff8a3d;'>Full Warranty Report →</a>")

    # ══ 6. NAVIGATION ══
    elif match(['where is','how to go','navigate','where can i']):
        nav_map = {
            'asset': ("<a href='/assets' style='color:#ff8a3d;'>📦 Assets</a>", "view and manage all assets"),
            'ticket': ("<a href='/tickets' style='color:#ff8a3d;'>🎫 Tickets</a>", "manage support tickets"),
            'employee': ("<a href='/employees' style='color:#ff8a3d;'>👥 Employees</a>", "manage the staff directory"),
            'report': ("<a href='/reports' style='color:#ff8a3d;'>🖊️ Reports</a>", "access inventory, warranty, and department reports"),
            'analytic': ("<a href='/analytics' style='color:#ff8a3d;'>📊 Analytics</a>", "view charts and trends"),
            'audit': ("<a href='/audit-logs' style='color:#ff8a3d;'>📋 Audit Logs</a>", "see all create/update/delete actions"),
        }
        found = False
        for key, (link, desc) in nav_map.items():
            if key in user_msg:
                reply = f"Head over to {link} from the sidebar to {desc}."
                found = True; break
        if not found:
            reply = "I can guide you! Tell me where you need to go — Assets, Tickets, Employees, Reports, Analytics, or Audit Logs?"

    elif match(['add asset','new asset','register asset']):
        reply = "To add a new asset 📦, go to <a href='/add-asset' style='color:#ff8a3d;'>Add Asset</a>. Fill in the serial number, category, location, and optionally assign to an employee." if role == 'admin' else "Only admin users can add assets. Please contact your administrator."

    elif match(['export','download','excel']):
        reply = "You can export assets to Excel via <a href='/export-assets' style='color:#ff8a3d;'>Export Assets</a>, or export any report using the <b>'Export Excel'</b> button on the <a href='/reports' style='color:#ff8a3d;'>Reports page</a>."

    elif match(['qr','qr code','scan','barcode']):
        reply = "Every asset has an auto-generated QR code 📱 Open any asset's detail page to find it. It encodes the asset ID, name, and serial number for quick scanning."

    elif match(['dark mode','light mode','theme']):
        reply = "Toggle between 🌙 dark and ☀️ light mode using the moon/sun icon in the top-right corner. Your preference is saved automatically!"

    # ══ 7. QUICK ANALYTICS ══
    elif match(['analytic','statistics','insight','overview','summary','dashboard data']):
        total, in_store, deployed, maint, total_val = get_asset_stats()
        t_total, t_open, _, _, _, t_overdue = get_ticket_stats()
        e_total, _ = get_employee_stats()
        reply = (f"📊 <b>AAI IMS Quick Overview</b><br><br>"
                 f"<b>Assets:</b> {total} total | {deployed} deployed | {in_store} in store<br>"
                 f"<b>Inventory Value:</b> {fmt_currency(total_val)}<br>"
                 f"<b>Tickets:</b> {t_total} total | {t_open} open | ⚠️ {t_overdue} overdue<br>"
                 f"<b>Employees:</b> {e_total}<br><br>"
                 f"<a href='/analytics' style='color:#ff8a3d;'>Full Analytics →</a>")

    # ══ 8. ROLE-AWARE ══
    elif match(['my asset','my assigned asset','what asset do i have']):
        if role == 'employee':
            emp_id = session.get('employee_id')
            assets = Asset.query.filter_by(assigned_employee_id=emp_id).all()
            if assets:
                lines = '<br>'.join(f"• <b>{a.asset_name}</b> ({a.category}) — {a.status}" for a in assets)
                reply = f"You have <b>{len(assets)} asset(s)</b> assigned 📦<br><br>{lines}"
            else:
                reply = "You don't have any assets assigned to you right now. Contact your admin if you need one."
        else:
            reply = "As admin, you manage all assets. Check <a href='/assets' style='color:#ff8a3d;'>Assets →</a>"

    elif match(['my ticket','my raised ticket','tickets i raised']):
        if role == 'employee':
            emp_id = session.get('employee_id')
            tickets = Ticket.query.filter_by(raised_by_employee_id=emp_id).order_by(Ticket.raised_at.desc()).limit(5).all()
            if tickets:
                lines = '<br>'.join(f"• [{t.ticket_number}] {t.title} — <b>{t.status}</b>" for t in tickets)
                reply = f"Your recent tickets 🎫<br><br>{lines}"
            else:
                reply = "You haven't raised any tickets yet. If you have an issue, click 'Raise Ticket' on your dashboard."
        else:
            reply = "As admin, view all tickets at <a href='/tickets' style='color:#ff8a3d;'>Tickets →</a>"

    # ══ 9. SMART FALLBACK ══
    else:
        if any(k in user_msg for k in ['asset','inventory']):
            total, in_store, deployed, _, _ = get_asset_stats()
            reply = f"I caught 'asset' in your question! We have <b>{total} assets</b> — {deployed} deployed, {in_store} in store. Try: <i>'show asset summary'</i> or <i>'total asset value'</i>"
        elif any(k in user_msg for k in ['ticket','issue','problem','complaint']):
            _, open_t, _, _, _, overdue = get_ticket_stats()
            reply = f"Sounds ticket-related! Currently <b>{open_t} open tickets</b>, <b>{overdue} overdue</b>. Try: <i>'show ticket summary'</i>"
        elif any(k in user_msg for k in ['employee','staff','person','user']):
            total, _ = get_employee_stats()
            reply = f"Sounds employee-related! We have <b>{total} employees</b>. Try: <i>'who has the most assets?'</i>"
        else:
            fallbacks = [
                "Hmm, I'm not sure I understand that one. Try asking about assets, tickets, employees, or costs! 🤔",
                "That's a bit outside my current knowledge. I'm best at inventory, tickets, and staff queries. 😊",
                "I didn't quite catch that! Try: <i>'How many open tickets are there?'</i> or <i>'Show asset summary'</i>",
                "I'm still learning! For now I handle asset tracking, ticket management, and employee info. What would you like to know? 🤖"
            ]
            reply = rnd.choice(fallbacks)

    return jsonify({'reply': reply})

# ===================== MAIN =====================

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        admin = User.query.filter_by(role='admin').first()
        if not admin:
            admin = User(username='AAI_atharv', password='#atharvintern07', role='admin')
            db.session.add(admin)
            db.session.commit()
            print('✅ Admin created: AAI_atharv / #atharvintern07')
    socketio.run(app, debug=True, host='127.0.0.1', port=5000)
